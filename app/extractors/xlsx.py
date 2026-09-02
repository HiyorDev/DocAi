from openpyxl import load_workbook

def extract_xlsx(path: str) -> str:

    workbook = load_workbook(path)

    text = ""

    for sheet in workbook.worksheets:

        text += f"Hoja {sheet.title}\n "

        for row in sheet.iter_rows(values_only =True):

            values = {
                str(value)
                for value in row
                if value is not None
            }

            text += " | ".join(values) + "\n"


    return text